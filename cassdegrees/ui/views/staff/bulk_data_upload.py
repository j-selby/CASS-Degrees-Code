import csv
from io import TextIOWrapper
import openpyxl

from api.models import CourseModel, SubplanModel
from django.contrib.auth.decorators import login_required
from django.shortcuts import render


# Submit file with courses or subplans with the following formats:
# Note that column order does not matter, as long as data corresponds to the order of the first row.

# Courses:
# code%name%units%offeredYears%offeredSem1%offeredSem2%offeredSummer%offeredAutumn%offeredWinter%offeredSpring%otherOffering
# ARTS1001%Introduction to Arts%6%ALL%True%False%True%True%True%True%False
# ...

# Subplans:
# code%year%name%units%planType
# ARTI-SPEC%2016%Artificial Intelligence%24%SPEC
# ...

# Support is available for excel sheets in the same rules as above (no % needed, but put each word between
# the % sign in a new cell in that row).

# Support is available for CASS' custom teaching plan excel file.
import os


@login_required
def bulk_data_upload(request):
    context = dict()
    context['upload_type'] = ['Courses', 'Subplans']
    content_type = request.GET.get('type')

    if content_type in context['upload_type']:
        context['current_tab'] = content_type

    if request.method == 'POST':
        base_model_url = request.build_absolute_uri('/api/model/')

        # Open file in text mode:
        # https://stackoverflow.com/questions/16243023/how-to-resolve-iterator-should-return-strings-not-bytes
        raw_uploaded_file = request.FILES['uploaded_file']
        file_type = os.path.splitext(str(raw_uploaded_file))[1]

        # Put the supported file extensions in this list
        supported_file_types = [".xlsx", ".xls", ".csv"]

        if file_type not in supported_file_types:
            context['user_msg'] = "Failed to upload file... " \
                                  "File format '" + file_type + "' is not supported! <br>" \
                                  "Please make sure the file extension and contents are correct."
            context['err_type'] = "error"
            return render(request, 'staff/bulkupload.html', context=context)

        uploaded_file = TextIOWrapper(raw_uploaded_file, encoding=request.encoding)

        # First row contains the column type headings (code, name etc). We can't add them to the db.
        first_row_checked = False

        # Check if any errors or successes appear when uploading the files.
        # Used for determining type of message to show to the user on the progress of their file upload.
        any_error = False
        any_success = False
        failed_to_upload = []
        correctly_uploaded = []

        # If the uploaded file was an excel sheet, convert it to a format that is the same as the output of when
        # the percent separated value files are processed by the else statement below.
        # This is done so that the code below the if else statement here works for both types of files.
        if uploaded_file.name[-4:] == "xlsx" or uploaded_file.name[-3:] == "xls":
            # Used https://www.pythoncircle.com/post/591/how-to-upload-and-process-the-excel-file-in-django/
            # to help me read the excel files.
            excel_file = openpyxl.load_workbook(raw_uploaded_file)
            sheet = excel_file["Sheet1"]

            # This is not the best way to make sure
            cass_course_custom_format = (list(sheet.iter_rows())[0][0].value is None)

            uploaded_file = list()

            # If the uploaded excel sheet is in format of what CASS already has, do special processing.
            # Change this if statement when the checkbox is implemented.
            if cass_course_custom_format:
                columns_set = False
                col_index = dict()

                # Add columns that are matching the custom format designed in this script, detailed in line 10 - 26.
                uploaded_file.append([
                    "code",
                    "name",
                    "units",
                    "offeredYears",
                    "offeredSem1",
                    "offeredSem2",
                    "offeredSummer",
                    "offeredAutumn",
                    "offeredWinter",
                    "offeredSpring",
                    "otherOffering"
                ])
                for row in sheet.iter_rows():
                    # This is the initialisation phase, where the file reader determines the column positions.
                    col_counter = 0
                    if not columns_set:
                        for cell in row:
                            # Determine the column positions of the excel sheet.
                            if not columns_set:
                                if cell.value is not None:
                                    # Find and store the index of all the column positions
                                    col_index[cell.value] = col_counter
                                    col_counter += 1

                                # If every column has been indexed, then mark columns_set as true.
                                if col_counter == len(row):
                                    columns_set = True

                    # Once the column positions are set,
                    # check the offerings specified in the 'Semester' and 'Sessions' columns and process accordingly.
                    else:
                        code = row[col_index["Course Code"]].value
                        name = row[col_index["Course Title"]].value

                        # Made all cases to be upper to make comparisons easier
                        semesters = str(row[col_index["Semesters"]].value)
                        sessions = str(row[col_index["Sessions"]].value)
                        comments = str(row[col_index["Comments"]].value)

                        semesters = semesters.upper() if semesters is not None else ""
                        sessions = sessions.upper() if sessions is not None else ""
                        comments = comments.upper() if comments is not None else ""

                        # Semester columns with 'other' and non-empty Comments column are marked with 'other_offering'.
                        other_offering = semesters == "OTHER" or comments != ""

                        sem_offer_set = False  # Marks whether semester 1, 2 offerings are set in the "Semesters" column

                        # Skip all courses that is marked for disestablishment.
                        if "DISESTABLISH" in comments:
                            skipped_course = "{} - {} - Marked for Disestablishment".format(code, name)

                            failed_to_upload.append(skipped_course)
                            any_error = True
                            continue

                        # If no semester offering details are mentioned, then initially set the following variables
                        # as follows, but can be overwritten later based on what the 'Sessions' column contains.
                        if semesters == "OTHER" or semesters == "OFFERED ONLY IN SESSIONS":
                            years_offered = "OTHER"
                            s1_offer = False
                            s2_offer = False

                        # If information on semester offerings are available, then extract all relevant information.
                        else:
                            if "EVERY" in semesters:
                                years_offered = "ALL"
                            elif "EVEN" in semesters:
                                years_offered = "EVEN"
                            elif "ODD" in semesters:
                                years_offered = "ODD"
                            else:
                                years_offered = "OTHER"

                            if semesters == "EVERY SEMESTER":
                                s1_offer = True
                                s2_offer = True
                            else:
                                s1_offer = "S1" in semesters
                                s2_offer = "S2" in semesters

                            sem_offer_set = True

                        # All rows that contains "all" in the "Sessions" column has referred to all sessions.
                        if "ALL" in sessions:
                            sum_offer = True
                            aut_offer = True
                            win_offer = True
                            spr_offer = True

                        # If individual sessions are mentioned, then extract this info by string comparison.
                        else:
                            sum_offer = "SUMMER" in sessions or "SUMMER" in comments
                            aut_offer = "AUTUMN" in sessions or "AUTUMN" in comments
                            win_offer = "WINTER" in sessions or "WINTER" in comments
                            spr_offer = "SPRING" in sessions or "SPRING" in comments

                        # If there are no information on offered years (odd, even etc) in the "Semesters" column,
                        # Check if there are any in the "Sessions" column and overwrite the years_offered.
                        # Also, there are no courses where offered years for semesters and sessions are different,
                        # so the overwrite is acceptable.
                        if "EVERY" in sessions:
                            years_offered = "ALL"
                        elif "EVEN" in sessions:
                            years_offered = "EVEN"
                        elif "ODD" in sessions:
                            years_offered = "ODD"

                        # If semesters aren't already set at this point, check the "Sessions" and "Comments" column to
                        # see if there are any information on semester offerings (e.g. S1 in 2020 only).
                        if not sem_offer_set:
                            if "SEMESTERS" in sessions or "EVERY SEMESTER" in sessions:
                                s1_offer = True
                                s2_offer = True
                            else:
                                s1_offer = "S1" in sessions or "SEMESTER 1" in sessions or \
                                           "S1" in comments or "SEMESTER 1" in comments
                                s2_offer = "S2" in sessions or "SEMESTER 2" in sessions or \
                                           "S2" in comments or "SEMESTER 2" in comments

                        # Assume all courses are worth 6 units,
                        # since unit value is not included in the CASS course list excel sheet.
                        row_data = [code, name, 6, years_offered, s1_offer, s2_offer,
                                    sum_offer, aut_offer, win_offer, spr_offer, other_offering]
                        uploaded_file.append(row_data)

            # If the uploaded excel sheet is in custom format specified in line 23,
            # then simply convert from the excel format to a list of lists.
            else:
                for row in sheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    uploaded_file.append(row_data)

        else:
            # Reading the '%' using the csv import module came from:
            # https://stackoverflow.com/questions/13992971/reading-and-parsing-a-tsv-file-then-manipulating-it-for-saving-as-csv-efficie

            # % is used instead of comma since the course name may include commas (which would break this function)
            uploaded_file = csv.reader(uploaded_file, delimiter='%')

        # Stores the index of the column containing the data type of each row,
        # so that the right data is stored in the right column
        # This would also allow columns to be in any order, and courses/subplans would still be added.
        map = dict()
        for row in uploaded_file:
            if first_row_checked:
                if content_type == 'Courses':
                    # If number of columns from file doesn't match the model, return error to user.
                    if len(row) != 11:
                        any_error = True
                        break

                    course_instance = CourseModel()
                    course_instance.code = row[map['code']]
                    course_instance.name = row[map['name']]
                    course_instance.units = int(row[map['units']])

                    course_instance.offeredYears = row[map['offeredYears']].upper()

                    # str() is wrapped around the value because if CASS supported file gets uploaded,
                    # then boolean values would be added to the rows instead of string "True" and "False".
                    # This is to support both string and boolean versions of True and False.
                    course_instance.offeredSem1 = str(row[map['offeredSem1']]).upper() == "TRUE"
                    course_instance.offeredSem2 = str(row[map['offeredSem2']]).upper() == "TRUE"

                    course_instance.offeredSummer = str(row[map['offeredSummer']]).upper() == "TRUE"
                    course_instance.offeredAutumn = str(row[map['offeredAutumn']]).upper() == "TRUE"
                    course_instance.offeredWinter = str(row[map['offeredWinter']]).upper() == "TRUE"
                    course_instance.offeredSpring = str(row[map['offeredSpring']]).upper() == "TRUE"

                    course_instance.otherOffering = str(row[map['otherOffering']]).upper() == "TRUE"

                    # It is assumed that all courses that are added are active.
                    course_instance.currentlyActive = True

                    course_str = course_instance.code + " - " + course_instance.name

                    # Save the course instance
                    try:
                        course_instance.save()
                        any_success = True
                        correctly_uploaded.append(course_str)
                    except:
                        error_message = course_str + " Couldn't add: Check for duplicate course"
                        failed_to_upload.append(error_message)
                        any_error = True

                elif content_type == 'Subplans':
                    if len(row) != 5:
                        any_error = True
                        break

                    subplan_instance = SubplanModel()
                    subplan_instance.code = row[map['code']]
                    subplan_instance.year = int(row[map['year']])
                    subplan_instance.name = row[map['name']]
                    subplan_instance.units = int(row[map['units']])
                    subplan_instance.planType = str(row[map['planType']])
                    subplan_str = str(subplan_instance.year) + " - " + subplan_instance.code + " - " + \
                        subplan_instance.name

                    # Save the subplan instance
                    try:
                        subplan_instance.save()
                        any_success = True
                        correctly_uploaded.append(subplan_str)
                    except:
                        any_error = True
                        failed_to_upload.append(subplan_str)

            else:
                i = 0
                for col in row:
                    map[col] = i
                    i += 1
                first_row_checked = True

        # If there are too many items to display, then only display the first 5 and numerically show how many more have
        # succeeded/failed.
        show_count = 5
        upload_count = len(correctly_uploaded)
        if upload_count > show_count:
            correctly_uploaded = correctly_uploaded[0:show_count]
            correctly_uploaded.append("... and {} more items".format(upload_count - show_count))

        fail_count = len(failed_to_upload)
        if fail_count > show_count:
            failed_to_upload = failed_to_upload[0:show_count]
            failed_to_upload.append("... and {} more items".format(fail_count - show_count))

        # Display error messages depending on the level of success of bulk upload.
        # There are 3 categories: All successful, some successful or none successful.
        if any_success and not any_error:
            context['user_msg'] = "All items has been added successfully!"
            context['err_type'] = "success"

        elif any_success and any_error:
            # for course in failed_to_upload:
            failed_str = ""
            correct_str = ""

            for record in failed_to_upload:
                failed_str = failed_str + "<br> - " + record

            for record in correctly_uploaded:
                correct_str = correct_str + "<br> - " + record

            context['user_msg'] = "Some items could not be added. They may already be present in the database" \
                                  " or the data may not be in the correct format. Please check the " \
                                  + content_type + \
                                  " list and try manually adding ones that failed through the dedicated " \
                                  "forms. <br><br>The following " + content_type + " could not be added: " + \
                                  failed_str + "<br><br> " \
                                  "The following " + content_type + " uploaded successfully: " + correct_str
            context['err_type'] = "warn"

        elif not any_success and any_error:
            context['user_msg'] = "All items failed to be added. " \
                                  "Either you have already uploaded the same contents, " \
                                  "or the format of the file is incorrect. Please try again."
            context['err_type'] = "error"

    return render(request, 'staff/bulkupload.html', context=context)
